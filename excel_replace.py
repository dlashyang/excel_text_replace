import sys
import openpyxl
import json


def create_cfg_example():
    cfg = {
        "source_file": "sample.xlsx",
        "target_file": "target.xlsx",
        "sheet": "Sheet2",
        "mode": "basic/RE"
    }

    cfg["replace"] = []
    cfg["replace"].append({"text_original/pattern": "", "text_new": ""})
    cfg["replace"].append({"text_original/pattern": "", "text_new": ""})

    with open("cfg.json.example", "w+") as f:
        json.dump(cfg, f, indent=4)


def main(src_file=None, dst_file=None):
    with open("cfg.json", "r") as f:
        cfg = json.load(f)

    if ((src_file is None) or (dst_file is None)):
        (src_file, dst_file) = (cfg["source_file"], cfg["target_file"])

    (sheet_name, mode) = (cfg["sheet"], cfg["mode"])
    print(f"mode: {mode}: Currently only simple mode is supported")
    replace_set = cfg["replace"]

    wb = openpyxl.load_workbook(src_file)
    ws = wb[sheet_name]

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            s = ws.cell(r, c).value
            if ((s is None) or (isinstance(s, str))):
                continue

            for replace in replace_set:
                (text_old, text_new) = (replace["text_original/pattern"],
                                        replace["text_new"])
                if text_old in s:
                    s = s.replace(text_old, text_new)
                    ws.cell(r, c).value = s
                    print(f"row {r} col {c} updated: {text_old} -> {text_new}")

    wb.save(dst_file)


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        (input_file, target_file) = sys.argv[-2:]
        main(input_file, target_file)
    else:
        main()
        # create_cfg_example()
